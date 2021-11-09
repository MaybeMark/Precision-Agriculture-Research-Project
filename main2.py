from statistics import pstdev
import pandas as pd
import numpy as np
import openpyxl
import random

matrix1 = pd.DataFrame(np.around(np.random.uniform(0, 10, size=(10, 10)), decimals=3), columns=list('ABCDEFGHIJ'))
# ^ Creates a 10X10 Dataframe with random floating point numbers between the range of 0-10
file = 'Initial.xlsx'  # Adds dataframe to an excel sheet
matrix2 = matrix1.copy()
with pd.ExcelWriter(file) as writer:  # Adds dataframe to an excel sheet
    matrix1.to_excel(writer, sheet_name='Sheet 1')
    # matrix2.to_excel(writer, sheet_name='Sheet 2')

stack = matrix1.stack()  # Flattens Dataframe (makes it easier to perform statistical operations on the data)
mean3 = round(stack.mean(), 4)  # Mean of entire Dataframe
stdev_s = round(stack.std(), 4)  # Sample Standard Deviation
stdev_p = round(pstdev(stack), 4)  # Population Standard Deviation

print(matrix1)
print("Mean: "+str(mean3))
print("Sample Standard Deviation: "+str(stdev_s))
print("Population Standard Deviation: "+str(stdev_p))


def randpick():
    l = 0
    wb = openpyxl.load_workbook("C:\\Users\\markd.LAPTOP-UMFS8BI9\\PycharmProjects\\USDA\\Initial.xlsx")
    ws = wb.active
    wrksht1 = wb['Sheet 1']
    # wrksht2 = wb['Sheet 2']
    while l < 20:
        randcell = [random.randint(2, 11), random.randint(2, 11)]
        print(randcell)
        l += 1
        c = wrksht1.cell(randcell[0], randcell[1]).value
        if randcell[1] != 2:
            c2 = wrksht1.cell(randcell[0], (randcell[1]-1)).value
            if c > 5:
                newc = c + random.randint(1, 2)
                wrksht1.cell(randcell[0], randcell[1]).value = newc

            elif c < 5:
                newc = c + c2
                wrksht1.cell(row=randcell[0], column=randcell[1]).value = newc
                wrksht1.cell(row=randcell[0], column=randcell[1] - 1).value = 0

        else:
            continue

    wb.save("New.xlsx")


randpick()
matrix2 = pd.read_excel('New.xlsx', sheet_name = 'Sheet 1', engine='openpyxl')
print(matrix2)
matrix3 = matrix2.subtract(matrix1)
matrix3.to_excel("Difference.xlsx")
print(matrix3)
