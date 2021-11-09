import openpyxl
import random

# Step 1
# ----------------------------------------------------------------------------------------------------------------------
wb = openpyxl.load_workbook(
    "C:\\Users\\markd.LAPTOP-UMFS8BI9\\PycharmProjects\\USDA\\APU_5_2017_soybeans_clip (1).xlsx")
sheet = wb.worksheets[0]
randcell1 = [random.randint(7, sheet.max_row), random.randint(0, sheet.max_column)]
randcell2 = [random.randint(7, sheet.max_row), random.randint(0, sheet.max_column)]
y1 = sheet.cell(randcell1[0], randcell1[1]).value
y2 = sheet.cell(randcell2[0], randcell2[1]).value
l = 0

# To make sure the two random cells are not the same
while l < 100:
    if randcell1 == randcell2:
        randcell2 = [random.randint(7, sheet.max_row), random.randint(0, sheet.max_column)]
        y2 = sheet.cell(randcell2[0], randcell2[1]).value
    else:
        break

x = 1
m1 = (y1 - 50) / x
m2 = (y2 - 50) / x
print(randcell1)
print(randcell2)
print("y1: "+str(y1))
print("y2: "+str(y2))
print(m1)
print(m2)


# Step 2
# ----------------------------------------------------------------------------------------------------------------------
S0 = y1 + y2

# Step 3
# ----------------------------------------------------------------------------------------------------------------------
y1 = 50+(m1*0)
y2 = 50+(m2*2)
print("**Moved all fertilizer to y(p2)**")
print("New y(p1): "+str(y1))
print("New y(p2): "+str(y2))

# Step 4
# ----------------------------------------------------------------------------------------------------------------------
S1 = y1+y2
print("S1: "+str(S1))

# Step 5
# ----------------------------------------------------------------------------------------------------------------------
y2 = 50+(m2*0)
y1 = 50+(m1*2)
print("**Moved all fertilizer to y(p1)**")
print("New y(p1): "+str(y1))
print("New y(p2): "+str(y2))

S2 = y1+y2
print("S2: "+str(S2))

# Step 6
# ----------------------------------------------------------------------------------------------------------------------
if S0>S1 and S0>S2:
    print("S0 is the greatest sum")
elif S1>S0 and S1>S2:
    print("S1 is the greatest sum")
else:
    print("S2 is the greatest sum")
