import random
import math
import xlsxwriter
import openpyxl
import pandas

# Might use these (probably not)
# from openpyxl import load_workbook
# from openpyxl import cell
# -----------------------------------

row1 = []
row2 = []
row3 = []
row4 = []
row5 = []
row6 = []
row7 = []
row8 = []
row9 = []
row10 = []
rowall = []


# Creates in stores random values in the lists above

def randm():
    l = 0
    while l < 10:
        num = round(random.uniform(0, 10), 3)
        row1.append(num)
        l += 1

    l = 0
    while l < 10:
        num = round(random.uniform(0, 10), 3)
        row2.append(num)
        l += 1

    l = 0
    while l < 10:
        num = round(random.uniform(0, 10), 3)
        row3.append(num)
        l += 1

    l = 0
    while l < 10:
        num = round(random.uniform(0, 10), 3)
        row4.append(num)
        l += 1

    l = 0
    while l < 10:
        num = round(random.uniform(0, 10), 3)
        row5.append(num)
        l += 1

    l = 0
    while l < 10:
        num = round(random.uniform(0, 10), 3)
        row6.append(num)
        l += 1

    l = 0
    while l < 10:
        num = round(random.uniform(0, 10), 3)
        row7.append(num)
        l += 1

    l = 0
    while l < 10:
        num = round(random.uniform(0, 10), 3)
        row8.append(num)
        l += 1

    l = 0
    while l < 10:
        num = round(random.uniform(0, 10), 3)
        row9.append(num)
        l += 1

    l = 0
    while l < 10:
        num = round(random.uniform(0, 10), 3)
        row10.append(num)
        l += 1


# Creates an Excel Sheet and stores values in 10X10 matrix
def write():
    workbook = xlsxwriter.Workbook('Initial.xlsx')
    worksheet = workbook.add_worksheet()
    # worksheet.print_area('A1:J1')
    # worksheet.print_area('A1:A10')
    #
    #
    row = 0
    column = 0
    for item in row1:
        worksheet.write(row, column, item, )
        column += 1

    row += 1
    column = 0
    for item in row2:
        worksheet.write(row, column, item, )
        column = column + 1

    row += 1
    column = 0
    for item in row3:
        worksheet.write(row, column, item, )
        column = column + 1

    row += 1
    column = 0
    for item in row4:
        worksheet.write(row, column, item, )
        column = column + 1

    row += 1
    column = 0
    for item in row5:
        worksheet.write(row, column, item, )
        column = column + 1

    row += 1
    column = 0
    for item in row6:
        worksheet.write(row, column, item, )
        column = column + 1

    row += 1
    column = 0
    for item in row7:
        worksheet.write(row, column, item, )
        column = column + 1

    row += 1
    column = 0
    for item in row8:
        worksheet.write(row, column, item, )
        column = column + 1

    row += 1
    column = 0
    for item in row9:
        worksheet.write(row, column, item, )
        column = column + 1

    row += 1
    column = 0
    for item in row10:
        worksheet.write(row, column, item, )
        column = column + 1

    workbook.close()


# Calculates mean and standard deviation
def stats():
    sum1 = 0
    for item in range(0, len(row1)):
        sum1 = sum1 + row1[item]

    sum2 = 0
    for item in range(0, len(row2)):
        sum2 = sum2 + row2[item]
    sum3 = 0
    for item in range(0, len(row3)):
        sum3 = sum3 + row3[item]
    sum4 = 0
    for item in range(0, len(row4)):
        sum4 = sum4 + row4[item]
    sum5 = 0
    for item in range(0, len(row5)):
        sum5 = sum5 + row5[item]
    sum6 = 0
    for item in range(0, len(row6)):
        sum6 = sum6 + row6[item]
    sum7 = 0
    for item in range(0, len(row7)):
        sum7 = sum7 + row7[item]
    sum8 = 0
    for item in range(0, len(row8)):
        sum8 = sum8 + row8[item]
    sum9 = 0
    for item in range(0, len(row9)):
        sum9 = sum9 + row9[item]
    sum10 = 0
    for item in range(0, len(row10)):
        sum10 = sum10 + row10[item]
    global mean
    mean = (sum1 + sum2 + sum3 + sum4 + sum5 + sum6 + sum7 + sum8 + sum9 + sum10) / 100
    print(round(mean, 4))

    # Standard Deviation Trial
    sum1b = 0
    for item in range(0, len(row1)):
        sum1b = sum1b + (row1[item] - mean) ** 2

    sum2b = 0
    for item in range(0, len(row2)):
        sum2b = sum2b + (row2[item] - mean) ** 2
    sum3b = 0
    for item in range(0, len(row3)):
        sum3b = sum3b + (row3[item] - mean) ** 2
    sum4b = 0
    for item in range(0, len(row4)):
        sum4b = sum4b + (row4[item] - mean) ** 2
    sum5b = 0
    for item in range(0, len(row5)):
        sum5b = sum5b + (row5[item] - mean) ** 2
    sum6b = 0
    for item in range(0, len(row6)):
        sum6b = sum6b + (row6[item] - mean) ** 2
    sum7b = 0
    for item in range(0, len(row7)):
        sum7b = sum7b + (row7[item] - mean) ** 2
    sum8b = 0
    for item in range(0, len(row8)):
        sum8b = sum8b + (row8[item] - mean) ** 2
    sum9b = 0
    for item in range(0, len(row9)):
        sum9b = sum9b + (row9[item] - mean) ** 2
    sum10b = 0
    for item in range(0, len(row10)):
        sum10b = sum10b + (row10[item] - mean) ** 2

    sumall = (sum1b + sum2b + sum3b + sum4b + sum5b + sum6b + sum7b + sum8b + sum9b + sum10b) / 100
    stndev = math.sqrt(sumall)
    print(round(stndev, 4))


# Picks 20 random cells in the Excel Sheet and complete required tasks
#
# 3. Write a short subroutine/module that picks a random cell in the 10 x 10 matrix created in Step 1
# 4. Execute the subroutine in Step 3 20 times doing the following:
#   a. For the random cell selected:
#           i.      If the value of that cell is greater than 5, change add a random number between 1 and 2 to the value already in that cell
#
#           ii.      If the value of that cell is less than 5,
#
#       1. Add the value from the cell in the same row, but one column to the left.
#       2. Set the value of the cell on column to the left to zero
#       3. If the randomly chosen cell is already in the left-most column, do not do 4.a.ii.1 or 4.a.ii.2
#
global randcell
def randpick():
    workbook2 = xlsxwriter.Workbook('New.xlsx')
    worksheet2 = workbook2.add_worksheet()
    l = 0
    while l < 20:
        randcell = [random.randint(1, 10), random.randint(1, 10)]
        print(randcell)
        l += 1
        path = 'C:\\Users\\markd.LAPTOP-UMFS8BI9\\PycharmProjects\\USDA\\Initial.xlsx'
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        c = sheet_obj.cell(row=randcell[0], column=randcell[1])
        if randcell[1] != 1:
            c2 = sheet_obj.cell(row=randcell[0], column=randcell[1] - 1)
            if c.value > 5:
                newc = c.value + random.randint(1, 2)

                if randcell[0] == 1:
                    row1[randcell[1] - 1] = newc

                elif randcell[0] == 2:
                    row2[randcell[1] - 1] = newc

                elif randcell[0] == 3:
                    row3[randcell[1] - 1] = newc

                elif randcell[0] == 4:
                    row4[randcell[1] - 1] = newc

                elif randcell[0] == 5:
                    row5[randcell[1] - 1] = newc

                elif randcell[0] == 6:
                    row6[randcell[1] - 1] = newc

                elif randcell[0] == 7:
                    row7[randcell[1] - 1] = newc

                elif randcell[0] == 8:
                    row8[randcell[1] - 1] = newc

                elif randcell[0] == 9:
                    row9[randcell[1] - 1] = newc

                elif randcell[0] == 10:
                    row10[randcell[1] - 1] = newc
            elif c.value < 5:
                newc = c.value + c2.value

                if randcell[0] == 1:
                    row1[randcell[1] - 1] = newc
                    row1[randcell[1] - 2] = 0

                elif randcell[0] == 2:
                    row2[randcell[1] - 1] = newc
                    row2[randcell[1] - 2] = 0

                elif randcell[0] == 3:
                    row3[randcell[1] - 1] = newc
                    row3[randcell[1] - 2] = 0

                elif randcell[0] == 4:
                    row4[randcell[1] - 1] = newc
                    row4[randcell[1] - 2] = 0

                elif randcell[0] == 5:
                    row5[randcell[1] - 1] = newc
                    row5[randcell[1] - 2] = 0

                elif randcell[0] == 6:
                    row6[randcell[1] - 1] = newc
                    row6[randcell[1] - 2] = 0

                elif randcell[0] == 7:
                    row7[randcell[1] - 1] = newc
                    row7[randcell[1] - 2] = 0

                elif randcell[0] == 8:
                    row8[randcell[1] - 1] = newc
                    row8[randcell[1] - 2] = 0

                elif randcell[0] == 9:
                    row9[randcell[1] - 1] = newc
                    row9[randcell[1] - 2] = 0

                elif randcell[0] == 10:
                    row10[randcell[1] - 1] = newc
                    row10[randcell[1] - 2] = 0
        else:
            continue
    row = 0
    column = 0
    for item in row1:
        worksheet2.write(row, column, item, )
        column += 1

    row += 1
    column = 0
    for item in row2:
        worksheet2.write(row, column, item, )
        column = column + 1

    row += 1
    column = 0
    for item in row3:
        worksheet2.write(row, column, item, )
        column = column + 1

    row += 1
    column = 0
    for item in row4:
        worksheet2.write(row, column, item, )
        column = column + 1

    row += 1
    column = 0
    for item in row5:
        worksheet2.write(row, column, item, )
        column = column + 1

    row += 1
    column = 0
    for item in row6:
        worksheet2.write(row, column, item, )
        column = column + 1

    row += 1
    column = 0
    for item in row7:
        worksheet2.write(row, column, item, )
        column = column + 1

    row += 1
    column = 0
    for item in row8:
        worksheet2.write(row, column, item, )
        column = column + 1

    row += 1
    column = 0
    for item in row9:
        worksheet2.write(row, column, item, )
        column = column + 1

    row += 1
    column = 0
    for item in row10:
        worksheet2.write(row, column, item, )
        column = column + 1

    workbook2.close()




def subtract():
    l=0



randm()
write()
stats()
randpick()
stats()  # Gets stats for new 10x10 matrix
# subtract()
