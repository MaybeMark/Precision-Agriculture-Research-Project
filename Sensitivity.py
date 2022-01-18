import openpyxl
import Yield2 as yd
import pandas as pd

wb = openpyxl.load_workbook("Adjusted.xlsx")
wrksht1 = wb.create_sheet("1", 0)
n1 = 0.0
fertlist = [0.2, 0.4, 0.6, 0.8, 1]  # Range of Fertilizer Prices for Sensitivity Analysis
cornlist = [3, 4, 5, 6, 8]  # Range of Corn Prices for Sensitivity Analysis
cols = 87

for j in range(5):
    cyv = (((cornlist[j]/56)/454)*1000000)  # Corn Yield Value
    cyv = round(cyv,1)
    print(cyv)
    for i in range(5):
        y = 1
        z = 1
        listuni = []
        n1 = fertlist[i]
        listuni = yd.quadoptimize(n1, cyv)
        for item in listuni:
            if y == cols + 1:
                z = z + 1
                y = 1
            wrksht1.cell(row=z, column=y).value = item
            # print(str(i) + " " + str(j) + " " + str(item))
            y = y + 1
        wb.save("Adjusted" + str(j) + str(i) + ".xlsx")
        # Create Excel sheet with sensitivity test results
        # name corresponds with loop count

tablelist = []
cornlist2 = [3,4,6,8]
for j in range(4):
    cyv = (((cornlist2[j] / 56) / 454) * 1000000)
    cyv = round(cyv, 1)
    for i in range(5):
        fert = fertlist[i]
        b = 0.073
        c = 0.0001689
        b = (b * cyv) - fert
        c = c * cyv
        deriv = b / (2 * c)
        tablelist.append(round(deriv, 0))

print("Table:")
print(str(tablelist[0])+" "+str(tablelist[1])+" "+str(tablelist[2])+" "+str(tablelist[3])+" "+str(tablelist[4]))
print(str(tablelist[5])+" "+str(tablelist[6])+" "+str(tablelist[7])+" "+str(tablelist[8])+" "+str(tablelist[9]))
print(str(tablelist[10])+" "+str(tablelist[11])+" "+str(tablelist[12])+" "+str(tablelist[13])+" "+str(tablelist[14]))
print(str(tablelist[15])+" "+str(tablelist[16])+" "+str(tablelist[17])+" "+str(tablelist[18])+" "+str(tablelist[19]))

data = [[tablelist[0], tablelist[1], tablelist[2], tablelist[3], tablelist[4]],
        [tablelist[5], tablelist[6], tablelist[7], tablelist[8], tablelist[9]],
        [tablelist[10], tablelist[11], tablelist[12], tablelist[13], tablelist[14]],
        [tablelist[15], tablelist[16], tablelist[17], tablelist[18], tablelist[19]]]

matrix = pd.DataFrame(data)
matrix.columns = [0.2, 0.4, 0.6, 0.8, 1]
matrix.index = ['3', '4', '6', '8']
matrix.to_excel("TableFile.xlsx")
