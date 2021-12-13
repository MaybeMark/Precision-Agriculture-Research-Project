#  This program assumes linearity of response to nitrogen fertilizer
import openpyxl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import statistics
import os

# Step 1
# ----------------------------------------------------------------------------------------------------------------------
matrix = pd.read_excel('APU_5_2017_soybeans_clip (1).xlsx', header=5, engine='openpyxl')
matrix.to_excel('Linear Optimization - APU_5_2017_soybeans_clip (1).xlsx')

if os.path.exists("Adjusted.xlsx"):
    os.remove("Adjusted.xlsx")

list1 = matrix.stack()  # Flattens Dataframe into Series
list1_0 = []
std = round(list1.std(), 4)
list2 = list1.sort_values()  # Sorts the values in ascending order
list2_0 = []
median = statistics.median(list2)
for item in list1:  # Creates a list without the NO DATA cells
    if item > 0:
        list1_0.append(item)

for item in list2:  # Creates numerically sorted list without the NO DATA cells
    if item > 0:
        list2_0.append(item)
mean1 = statistics.mean(list1_0)
median = statistics.median(list2_0)
print(mean1)

print(list1.to_string())
print("\n\n\n\n\n\n")
print(list2.to_string())
print("\n\n\n\n\n\n")

#  Linear Optimization
listapp = []  # List of optimized yield
listfert = []  # List of fertilizer prescription
list3 = np.array_split(list2, 2)  # Splits the list of numbers into two lists
sum = 0.0
for i in list3:
    for j in i:
        sum = sum + j

sum2_0 = 0
sum2_0_1 = 0
sum2_1 = 0
sum2_1_1 = 0
for item in list1:
    if median > item > 0:  # if the cell is not a NO DATA cell and it is less than the median
        y1 = item

        x = 1
        m1 = (y1 - 50) / x
        # print(m1)
        if m1 < 0:
            y1_0 = 50
        else:
            y1_0 = 50 + (m1 * 1)  # Fertilizer application not adjusted
        y1 = 50 + (m1 * 0)  # Fertilizer application adjusted
        print(y1)
        listapp.append(y1)  # Creates list of optimized yield
        listfert.append(0)  # Creates list of prescribed fertilizer treatment
        sum2_0 = sum2_0 + y1  # Sums all of the results (after process)
        sum2_0_1 = sum2_0_1 + y1_0  # Sums all of the results (before process)
    elif item >= median and item > 0:  # if the cell is not a NO DATA cell and it is greater than the median
        y2 = item
        x = 1
        m2 = (y2 - 50) / x
        # print(m2)
        if m2 < 0:  # if a value is less than 0, the yield will be 50
            y2_0 = 50
        else:
            y2_0 = 50 + (m2 * 1)  # Not adjusted fertilizer application
        y2 = 50 + (m2 * 2)  # Adjusted fertilizer application
        print(y2)
        listapp.append(y2)
        listfert.append(2)
        sum2_1 = sum2_1 + y2  # Sums all of the results (after process)
        sum2_1_1 = sum2_1_1 + y2_0  # Sums all of the results (before process)
    else:  # if the cell is a NO DATA cell then it is just appended with no optimization
        listapp.append(item)
        listfert.append(item)

adjsum = 0.0
for i in listapp:
    adjsum = adjsum + i

mean = round(adjsum / len(listapp))

print("Original sum:" + str(sum))
print("Adjusted sum: " + str(sum2_0_1 + sum2_1_1))
print("Sum after process: " + str(adjsum))
print("Mean after process: " + str(mean))

print(listapp)

list3 = list2.to_numpy()

k = 0
for item in list3:  # Replaces all values less than 0 with 0
    if item < 0:
        list3[k] = 0
    k += 1

print("Standard Deviation: " + str(np.std(list3)))
print("Median: " + str(median))

print(len(matrix.columns))
print(len(matrix))
workbook = openpyxl.Workbook()
workbook.save("AdjustedTemp.xlsx")

i = 1
j = 1
wb = openpyxl.load_workbook("AdjustedTemp.xlsx")
wrksht1 = wb.create_sheet("Optimized Yield", 0)
wrksht2 = wb.create_sheet("Optimized Fertilizer")
wrksht3 = wb.create_sheet("Optimized Fertilzer Quadratic")

cols = len(matrix.columns)
for item in listapp:
    if j == cols+1:
        i = i + 1
        j = 1
    wrksht1.cell(row=i, column=j).value = item
    j = j + 1

i = 1
j = 1
for item in listfert:
    if j == cols+1:
        i = i + 1
        j = 1
    wrksht2.cell(row=i, column=j).value = item
    j = j + 1

# ----------------------------------------------------------------------------------------------------------------------
# Quadratic-plus-plateau (not done)
# ----------------------------------------------------------------------------------------------------------------------

listquad = []
i = 1
j = 1
for item in list1:
    if item>0:
        s = item / mean1  # Scaling Factor
        #  print(s)
        a = 6
        b = 0.073
        c = 0.0001689
        a = a * s * 208.5          # -------------------------------------------------------------------
        b = (b * s * 208.5) - .55  # Each coefficient is multiplied by the conversion and scaling factor
        c = c * s * 208.5          # -------------------------------------------------------------------
        deriv = b / (2 * c)
        print(str(item) + " " + str(s) + " " + str(deriv))
        listquad.append(deriv)
    else:
        listquad.append(item)

for item in listquad:
    if j == cols+1:
        i = i + 1
        j = 1
    wrksht3.cell(row=i, column=j).value = item
    print(str(i) + " " + str(j) + " " + str(item))
    j = j + 1

wb.save("Adjusted.xlsx")

if os.path.exists("AdjustedTemp.xlsx"):
    os.remove("AdjustedTemp.xlsx")
else:
    print("The file has already been removed.")

print(len(matrix))
print(len(matrix.columns))
# Frequency Distribution of data
kwargs = dict(alpha=0.5, bins=100)
plt.hist(list3, **kwargs, color='g')  # list3 is an array of list2
plt.title("Yield Distribution")
plt.xlabel('Yield')
plt.ylabel('Frequency')
plt.axvline(mean, color='k', linestyle='dashed', linewidth=1)  # this line shows mean on the graph
plt.show()  # this creates the histogram

#  EONR is calculated by the yield increase multiplied by the price of corn. That product is then subtracted by the
#   cost of N
